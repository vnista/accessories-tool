import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill


st.set_page_config(page_title="Honda Accessories – Tool accessori", layout="centered")


# === 1. Caricamento master IT (dal repo GitHub) ===
@st.cache_data
def load_master():
    return pd.read_excel("Master_Accessories_Merged.xlsx")


# === 2a. Funzione conversione EN → IT ===
def merge_files_overwrite(df_en: pd.DataFrame, df_master: pd.DataFrame) -> pd.DataFrame:
    master_small = df_master[["PARTNUMBER", "DESCRIPTION", "REMARK", "GROUP", "MASTER IMAGE"]].copy()
    master_small = master_small.rename(
        columns={
            "DESCRIPTION": "DESCRIPTION_IT",
            "REMARK": "REMARK_IT",
            "GROUP": "GROUP_IT",
            "MASTER IMAGE": "MASTER_IMAGE_IT",
        }
    )

    merged = df_en.merge(master_small, on="PARTNUMBER", how="left")

    if "DESCRIPTION" in merged.columns:
        merged["DESCRIPTION"] = merged["DESCRIPTION_IT"].fillna(merged["DESCRIPTION"])
    if "REMARK" in merged.columns:
        merged["REMARK"] = merged["REMARK_IT"].fillna(merged["REMARK"])
    if "GROUP" in merged.columns:
        merged["GROUP"] = merged["GROUP_IT"].fillna(merged["GROUP"])
    if "MASTER IMAGE" in merged.columns:
        merged["MASTER IMAGE"] = merged["MASTER_IMAGE_IT"].fillna(merged["MASTER IMAGE"])

    merged["NOT_FOUND"] = merged["DESCRIPTION_IT"].isna()

    cols_finali = ["PARTNUMBER", "DESCRIPTION", "REMARK", "GROUP", "MASTER IMAGE", "NOT_FOUND"]
    cols_finali = [c for c in cols_finali if c in merged.columns]

    result = merged[cols_finali]
    result = apply_wheel_price(result)

    return result


# === 2b. Funzione pulizia file già in italiano ===
def process_italian_file(df_it: pd.DataFrame, df_master: pd.DataFrame) -> pd.DataFrame:
    mandatory = "Listino comprensivo di IVA, montaggio escluso"

    if "REMARK" in df_it.columns:
        df_it["REMARK"] = df_it["REMARK"].fillna("").astype(str).str.strip()

        def add_mandatory(text):
            if text == "":
                return mandatory
            if mandatory.lower() in text.lower():
                return text
            return f"{text} {mandatory}"

        df_it["REMARK"] = df_it["REMARK"].apply(add_mandatory)

    if "GROUP" in df_it.columns and "PARTNUMBER" in df_it.columns:
        gmap = (
            df_master[["PARTNUMBER", "GROUP"]]
            .dropna(subset=["PARTNUMBER"])
            .drop_duplicates(subset=["PARTNUMBER"], keep="first")
            .set_index("PARTNUMBER")["GROUP"]
        )
        df_it["GROUP"] = df_it["PARTNUMBER"].map(gmap).fillna(df_it["GROUP"])

    df_it = apply_wheel_price(df_it)

    return df_it


# === 2c. Moltiplicazione x4 prezzo per CERCHI IN LEGA ===
def apply_wheel_price(df: pd.DataFrame) -> pd.DataFrame:
    """
    Per tutte le righe dove GROUP == 'CERCHI IN LEGA',
    moltiplica il valore della colonna del prezzo IVA inclusa per 4.
    Cerca la colonna prezzo in modo flessibile (case-insensitive).
    """
    price_col = None
    for col in df.columns:
        col_clean = str(col).strip().upper()
        if "PRICE" in col_clean and "INCL" in col_clean and "VAT" in col_clean:
            price_col = col
            break

    if price_col is None or "GROUP" not in df.columns:
        st.warning(
            f"Colonna prezzo non trovata o colonna GROUP mancante.\n"
            f"Colonne disponibili nel file: {list(df.columns)}"
        )
        return df

    mask = df["GROUP"].astype(str).str.strip().str.upper() == "CERCHI IN LEGA"
    df[price_col] = pd.to_numeric(df[price_col], errors="coerce")
    df.loc[mask, price_col] = df.loc[mask, price_col] * 4

    return df


# === 3. Utility per creare il file Excel con evidenziazione giallo per CERCHI IN LEGA ===
def to_excel_download(df: pd.DataFrame) -> bytes:
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="ACCESSORIES")

        workbook = writer.book
        worksheet = writer.sheets["ACCESSORIES"]

        yellow_fill = PatternFill(
            start_color="FFFF00",
            end_color="FFFF00",
            fill_type="solid"
        )

        if "GROUP" in df.columns:
            for row_idx, group_val in enumerate(df["GROUP"], start=2):
                if str(group_val).strip().upper() == "CERCHI IN LEGA":
                    for col_idx in range(1, len(df.columns) + 1):
                        worksheet.cell(row=row_idx, column=col_idx).fill = yellow_fill

    return output.getvalue()


# === 4. Genera nome file di output ===
def build_output_filename(uploaded_filename: str) -> str:
    """
    Prende il nome del file originale (es. 'EU_25YM_HR-V_Hybrid_accessories.xlsx')
    e restituisce lo stesso nome con '_hondatool' aggiunto prima dell'estensione
    (es. 'EU_25YM_HR-V_Hybrid_accessories_hondatool.xlsx').
    """
    base = uploaded_filename.rsplit(".", 1)[0]
    return f"{base}_hondatool.xlsx"


# === 5. App Streamlit ===
def main():
    st.title("Honda Accessories – Tool accessori")

    mode = st.radio(
        "Seleziona funzione:",
        ["1) Conversione EN → IT", "2) Pulizia file IT (REMARK + GROUP)"],
        index=0,
    )

    # Carico il master IT una volta sola
    try:
        df_master = load_master()
        st.success(f"Database IT caricato: {len(df_master)} righe dal master.")
    except FileNotFoundError:
        st.error(
            "Impossibile trovare il file `Master_Accessories_Merged.xlsx`.\n\n"
            "Verifica che sia nel **repo GitHub**, nella **stessa cartella di app.py**, "
            "e che il nome sia esattamente identico (maiuscole, estensione)."
        )
        return

    # ───────────────────────────────────────────
    # MODALITÀ 1: Conversione EN → IT
    # ───────────────────────────────────────────
    if mode.startswith("1"):
        st.markdown(
            "### Conversione EN → IT\n"
            "Carica il **file master accessori in inglese**.\n\n"
            "- Le intestazioni devono essere alla **riga 10**.\n"
            "- L'app sostituirà in italiano `DESCRIPTION`, `REMARK`, `GROUP` e `MASTER IMAGE`.\n"
            "- Le righe con `GROUP = CERCHI IN LEGA` avranno il prezzo (`PRICE INCL.VAT EUR`) "
            "**moltiplicato x4** ed evidenziate in **giallo**.\n"
            "- L'output avrà lo stesso nome del file caricato + `_hondatool`."
        )

        uploaded_file = st.file_uploader(
            "Carica il file master accessori in inglese (.xlsx / .xls)",
            type=["xlsx", "xls"],
            key="upload_en",
        )

        if uploaded_file is not None:
            try:
                df_en = pd.read_excel(uploaded_file, header=9)
            except Exception as e:
                st.error(f"Errore nella lettura del file: {e}")
                return

            if "PARTNUMBER" not in df_en.columns:
                st.error(
                    "Il file (a partire dalla riga 10) deve avere una colonna chiamata **'PARTNUMBER'**."
                )
                return

            st.subheader("Anteprima file EN caricato")
            st.dataframe(df_en.head())

            if st.button("Esegui conversione EN → IT"):
                result = merge_files_overwrite(df_en, df_master)

                st.subheader("Anteprima risultato")
                st.dataframe(result.head())

                total = len(result)
                not_found = result["NOT_FOUND"].sum() if "NOT_FOUND" in result.columns else 0
                wheels = (
                    result["GROUP"].astype(str).str.strip().str.upper() == "CERCHI IN LEGA"
                ).sum() if "GROUP" in result.columns else 0

                st.info(
                    f"Righe totali: {total} – "
                    f"Codici NON trovati: {not_found} – "
                    f"Cerchi in lega (prezzo x4 + giallo): {wheels}"
                )

                excel_bytes = to_excel_download(result)
                output_filename = build_output_filename(uploaded_file.name)
                st.download_button(
                    label=f"Scarica {output_filename}",
                    data=excel_bytes,
                    file_name=output_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

    # ───────────────────────────────────────────
    # MODALITÀ 2: Pulizia file già in italiano
    # ───────────────────────────────────────────
    else:
        st.markdown(
            "### Pulizia file IT (REMARK + GROUP)\n"
            "Carica un file **già localizzato in italiano**.\n\n"
            "- Alla colonna `REMARK` verrà aggiunta la frase "
            "`\"Listino comprensivo di IVA, montaggio escluso\"`.\n"
            "- La colonna `GROUP` verrà riscritta con le voci italiane del database.\n"
            "- Le righe con `GROUP = CERCHI IN LEGA` avranno il prezzo (`PRICE INCL.VAT EUR`) "
            "**moltiplicato x4** ed evidenziate in **giallo**.\n"
            "- L'output avrà lo stesso nome del file caricato + `_hondatool`."
        )

        uploaded_file_it = st.file_uploader(
            "Carica il file accessori IT da aggiornare (.xlsx / .xls)",
            type=["xlsx", "xls"],
            key="upload_it",
        )

        if uploaded_file_it is not None:
            try:
                df_it = pd.read_excel(uploaded_file_it, header=9)
            except Exception as e:
                st.error(f"Errore nella lettura del file: {e}")
                return

            if "PARTNUMBER" not in df_it.columns:
                st.error("Il file deve avere una colonna **'PARTNUMBER'**.")
                return

            st.subheader("Anteprima file IT caricato")
            st.dataframe(df_it.head())

            if st.button("Esegui pulizia file IT"):
                result_it = process_italian_file(df_it, df_master)

                st.subheader("Anteprima risultato (REMARK + GROUP + prezzi cerchi aggiornati)")
                st.dataframe(result_it.head())

                wheels = (
                    result_it["GROUP"].astype(str).str.strip().str.upper() == "CERCHI IN LEGA"
                ).sum() if "GROUP" in result_it.columns else 0
                st.info(f"Cerchi in lega trovati (prezzo x4 + giallo applicato): {wheels}")

                excel_bytes_it = to_excel_download(result_it)
                output_filename_it = build_output_filename(uploaded_file_it.name)
                st.download_button(
                    label=f"Scarica {output_filename_it}",
                    data=excel_bytes_it,
                    file_name=output_filename_it,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )


if __name__ == "__main__":
    main()
